"""
Excel and PDF report generators.

These are pure functions: take the in-memory `current_analysis` dict and
return a BytesIO buffer ready to be sent with Flask `send_file`.
"""
from __future__ import annotations

import io
import traceback
from typing import Any

from flask import send_file

from . import analyzer
from .config import get_logger
from .utils import date_only, normalize_url

logger = get_logger(__name__)


# ══════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════
def build_excel_report() -> Any:
    """Return a Flask `send_file` response for the Excel report."""
    report = analyzer.current_analysis
    if not report or report.get("status") != "success":
        from flask import jsonify
        return jsonify({"error": "No completed report available"}), 404

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        aeo = report.get("aeo_scores", {})
        ins = report.get("insights", {})
        url = report.get("target_url", "")
        date = date_only(report.get("analyzed_at", ""))

        DARK, ACCENT = "0F1724", "4F8EF7"
        GREEN, RED, YELLOW = "22C55E", "EF4444", "F59E0B"
        WHITE, GRAY = "E2EAF5", "1E2D42"

        def thin_border():
            s = Side(style="thin", color="1E2D42")
            return Border(left=s, right=s, top=s, bottom=s)

        # ── Sheet 1: AEO Summary ──────────────────────────────────
        ws = wb.active
        ws.title = "AEO Summary"
        ws.sheet_view.showGridLines = False
        for col, w in zip("ABCDE", [28, 18, 18, 18, 28]):
            ws.column_dimensions[col].width = w

        ws.merge_cells("A1:E1")
        ws["A1"].value = f"AEO Report — {url}"
        ws["A1"].font = Font(bold=True, size=14, color=WHITE, name="Arial")
        ws["A1"].fill = PatternFill("solid", start_color=DARK)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:E2")
        ws["A2"].value = f"Generated: {date}  |  Scrapfly + Z.ai  |  Auto-runs: 1st · 15th · 30th"
        ws["A2"].font = Font(italic=True, size=10, color="8B949E", name="Arial")
        ws["A2"].fill = PatternFill("solid", start_color=DARK)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        for col, (label, bg) in enumerate([
            ("Metric", GRAY), ("Score", ACCENT),
            ("ChatGPT", "107040"), ("Gemini", "1A56A0"), ("Google AI", "8B5E00"),
        ], 1):
            c = ws.cell(row=4, column=col, value=label)
            c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        ws.row_dimensions[4].height = 24

        score_rows = [
            ("Overall AEO Score",    aeo.get("overall_aeo", 0),             "—", "—", "—"),
            ("ChatGPT Visibility",   aeo.get("chatgpt_visibility", 0),       aeo.get("chatgpt_visibility", 0), "—", "—"),
            ("Gemini Visibility",    aeo.get("gemini_visibility", 0),        "—", aeo.get("gemini_visibility", 0), "—"),
            ("Google AI Overview",   aeo.get("google_ai_overview", 0),       "—", "—", aeo.get("google_ai_overview", 0)),
            ("Content Quality",      aeo.get("content_quality", 0),          "—", "—", "—"),
            ("Structured Data",      aeo.get("structured_data", 0),          "—", "—", "—"),
            ("Conversational Ready", aeo.get("conversational_readiness", 0), "—", "—", "—"),
            ("Entity Clarity",       aeo.get("entity_clarity", 0),           "—", "—", "—"),
        ]
        for i, (label, *vals) in enumerate(score_rows, 5):
            ws.row_dimensions[i].height = 20
            bg_row = "131D2E" if i % 2 == 0 else "0F1724"
            c = ws.cell(row=i, column=1, value=label)
            c.font = Font(color=WHITE, name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg_row)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", indent=1)
            for j, v in enumerate(vals, 2):
                cell = ws.cell(row=i, column=j, value=v)
                cell.fill = PatternFill("solid", start_color=bg_row)
                cell.border = thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(v, int):
                    color = GREEN if v >= 70 else (YELLOW if v >= 40 else RED)
                    cell.font = Font(bold=True, color=color, name="Arial", size=11)
                else:
                    cell.font = Font(color="5A7299", name="Arial", size=10)

        row = len(score_rows) + 6
        for col, (title, items, bg) in enumerate([
            ("Strengths", aeo.get("strengths", []), GREEN),
            ("Weaknesses", aeo.get("weaknesses", []), RED),
        ], 1):
            c = ws.cell(row=row, column=col * 2 - 1, value=title)
            c.font = Font(bold=True, color=WHITE, size=11, name="Arial")
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=row, start_column=col*2-1, end_row=row, end_column=col*2)
            ws.row_dimensions[row].height = 22
            for k, item in enumerate(items, 1):
                r = row + k
                ws.row_dimensions[r].height = 18
                for col2 in range(col*2-1, col*2+1):
                    ws.cell(row=r, column=col2).fill = PatternFill("solid", start_color="131D2E")
                c2 = ws.cell(row=r, column=col*2-1, value=f"• {item}")
                c2.font = Font(color=WHITE, name="Arial", size=10)
                c2.alignment = Alignment(vertical="center", indent=1)
                ws.merge_cells(start_row=r, start_column=col*2-1, end_row=r, end_column=col*2)

        # ── Sheet 2: Competitors ──────────────────────────────────
        wc = wb.create_sheet("Competitors")
        wc.sheet_view.showGridLines = False
        for col, (w, label) in enumerate([
            (22, "Competitor"), (14, "AEO Score"), (30, "Strengths"),
            (30, "Weaknesses"), (30, "Opportunity"),
        ], 1):
            wc.column_dimensions[get_column_letter(col)].width = w
            c = wc.cell(row=1, column=col, value=label)
            c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
            c.fill = PatternFill("solid", start_color=ACCENT)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        wc.row_dimensions[1].height = 22

        for i, comp in enumerate(report.get("competitors", []), 2):
            bg = "131D2E" if i % 2 == 0 else "0F1724"
            wc.row_dimensions[i].height = 40
            for col, val in enumerate([
                comp.get("competitor_name", comp.get("competitor_url", "")),
                comp.get("their_aeo_score", ""),
                "\n".join(f"• {s}" for s in comp.get("strengths", [])),
                "\n".join(f"• {w}" for w in comp.get("weaknesses", [])),
                comp.get("opportunity", ""),
            ], 1):
                c = wc.cell(row=i, column=col, value=val)
                c.fill = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                c.border = thin_border()
                if col == 2 and isinstance(val, int):
                    color = GREEN if val >= 70 else (YELLOW if val >= 40 else RED)
                    c.font = Font(bold=True, color=color, name="Arial", size=12)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.font = Font(color=WHITE, name="Arial", size=10)

        # ── Sheet 3: FAQs & Recommendations ─────────────────────
        wf = wb.create_sheet("FAQs & Recommendations")
        wf.sheet_view.showGridLines = False
        wf.column_dimensions["A"].width = 40
        wf.column_dimensions["B"].width = 50
        wf.column_dimensions["C"].width = 14

        wf.merge_cells("A1:C1")
        wf["A1"].value = "Trending FAQs to Add to Your Site"
        wf["A1"].font = Font(bold=True, color=WHITE, size=12, name="Arial")
        wf["A1"].fill = PatternFill("solid", start_color=ACCENT)
        wf["A1"].alignment = Alignment(horizontal="center", vertical="center")
        wf.row_dimensions[1].height = 26

        for col, label in enumerate(["Question", "Answer", "AI Platform"], 1):
            c = wf.cell(row=2, column=col, value=label)
            c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
            c.fill = PatternFill("solid", start_color=GRAY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()

        for i, faq in enumerate(ins.get("trending_faqs", []), 3):
            bg = "131D2E" if i % 2 == 0 else "0F1724"
            wf.row_dimensions[i].height = 45
            for col, val in enumerate([faq.get("question", ""), faq.get("answer", ""), faq.get("ai_platform", "")], 1):
                c = wf.cell(row=i, column=col, value=val)
                c.font = Font(color=WHITE, name="Arial", size=10)
                c.fill = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                c.border = thin_border()

        rec_start = len(ins.get("trending_faqs", [])) + 5
        wf.merge_cells(f"A{rec_start}:C{rec_start}")
        wf[f"A{rec_start}"].value = "Strategic Recommendations"
        wf[f"A{rec_start}"].font = Font(bold=True, color=WHITE, size=12, name="Arial")
        wf[f"A{rec_start}"].fill = PatternFill("solid", start_color="8B3A3A")
        wf[f"A{rec_start}"].alignment = Alignment(horizontal="center", vertical="center")
        wf.row_dimensions[rec_start].height = 26

        for col, label in enumerate(["Category", "Action", "Priority"], 1):
            c = wf.cell(row=rec_start + 1, column=col, value=label)
            c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
            c.fill = PatternFill("solid", start_color=GRAY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()

        for i, rec in enumerate(ins.get("recommendations", []), rec_start + 2):
            bg = "131D2E" if i % 2 == 0 else "0F1724"
            p = rec.get("priority", "")
            pcol = GREEN if p == "LOW" else (YELLOW if p == "MEDIUM" else RED)
            wf.row_dimensions[i].height = 36
            for col, val in enumerate([rec.get("category", ""), rec.get("action", ""), p], 1):
                c = wf.cell(row=i, column=col, value=val)
                c.fill = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                c.border = thin_border()
                if col == 3:
                    c.font = Font(bold=True, color=pcol, name="Arial", size=10)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.font = Font(color=WHITE, name="Arial", size=10)

        # ── Sheet 4: 30-Day Plan ──────────────────────────────────
        wp = wb.create_sheet("30-Day Plan")
        wp.sheet_view.showGridLines = False
        wp.column_dimensions["A"].width = 16
        wp.column_dimensions["B"].width = 70

        wp.merge_cells("A1:B1")
        wp["A1"].value = "30-Day AEO Action Plan"
        wp["A1"].font = Font(bold=True, color=WHITE, size=13, name="Arial")
        wp["A1"].fill = PatternFill("solid", start_color=DARK)
        wp["A1"].alignment = Alignment(horizontal="center", vertical="center")
        wp.row_dimensions[1].height = 28

        plan_colors = [ACCENT, "107040", "1A56A0", "8B5E00"]
        for i, item in enumerate(ins.get("next_30_day_plan", []), 2):
            parts = item.split(":", 1)
            week = parts[0].strip()
            task = parts[1].strip() if len(parts) > 1 else item
            wp.row_dimensions[i].height = 36
            cw = wp.cell(row=i, column=1, value=week)
            cw.font = Font(bold=True, color=WHITE, name="Arial", size=11)
            cw.fill = PatternFill("solid", start_color=plan_colors[(i - 2) % 4])
            cw.alignment = Alignment(horizontal="center", vertical="center")
            cw.border = thin_border()
            ct = wp.cell(row=i, column=2, value=task)
            ct.font = Font(color=WHITE, name="Arial", size=11)
            ct.fill = PatternFill("solid", start_color="131D2E")
            ct.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
            ct.border = thin_border()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe = url.replace("https://", "").replace("/", "_")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"AEO_Report_{safe}_{date}.xlsx",
        )

    except Exception as exc:
        logger.error("Excel generation failed: %s", exc)
        from flask import jsonify
        return jsonify({"error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════
def build_pdf_report() -> Any:
    """Return a Flask `send_file` response for the PDF report."""
    report = analyzer.current_analysis
    if not report or report.get("status") != "success":
        from flask import jsonify
        return jsonify({"error": "No completed report available"}), 404

    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            HRFlowable, KeepTogether, Paragraph, SimpleDocTemplate,
            Spacer, Table, TableStyle,
        )

        aeo = report.get("aeo_scores", {})
        ins = report.get("insights", {})
        url = report.get("target_url", "")
        date = date_only(report.get("analyzed_at", ""))

        buf = io.BytesIO()
        w, h = A4
        content_w = w - 40 * mm
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=20*mm, bottomMargin=20*mm,
            leftMargin=20*mm, rightMargin=20*mm,
            title=f"AEO Report - {url}",
        )

        NAVY, BLUE, BLUE_LT = colors.HexColor("#1a2744"), colors.HexColor("#2563eb"), colors.HexColor("#dbeafe")
        GREEN, GREEN_LT = colors.HexColor("#16a34a"), colors.HexColor("#dcfce7")
        RED, RED_LT = colors.HexColor("#dc2626"), colors.HexColor("#fee2e2")
        AMBER, AMBER_LT = colors.HexColor("#d97706"), colors.HexColor("#fef3c7")
        GRAY_DK, GRAY_MD, GRAY_LT, GRAY_BD, WHITE = (
            colors.HexColor("#374151"), colors.HexColor("#6b7280"),
            colors.HexColor("#f3f4f6"), colors.HexColor("#e5e7eb"), colors.white,
        )

        def ps(name, **kw):
            return ParagraphStyle(name, **kw)

        S_SECTION = ps("sec", fontSize=13, fontName="Helvetica-Bold",
                       textColor=NAVY, spaceBefore=12, spaceAfter=5)
        S_FAQ_Q = ps("faqq", fontSize=10, fontName="Helvetica-Bold",
                     textColor=BLUE, spaceAfter=2, leading=14)
        S_FAQ_A = ps("faqa", fontSize=9, fontName="Helvetica",
                     textColor=GRAY_DK, spaceAfter=8, leading=13, leftIndent=12)
        S_FOOTER = ps("foot", fontSize=8, fontName="Helvetica",
                      textColor=GRAY_MD, alignment=TA_CENTER)

        def score_clr(v):
            try: v = int(v)
            except: return GRAY_MD
            return GREEN if v >= 70 else (AMBER if v >= 40 else RED)

        def score_bg(v):
            try: v = int(v)
            except: return GRAY_LT
            return GREEN_LT if v >= 70 else (AMBER_LT if v >= 40 else RED_LT)

        def rating(v):
            try: v = int(v)
            except: return "N/A"
            return "Excellent" if v >= 80 else "Good" if v >= 60 else "Fair" if v >= 40 else "Needs Work"

        def make_table(data, widths, extra=None):
            t = Table(data, colWidths=widths, repeatRows=1)
            base = [
                ("BACKGROUND",    (0,0), (-1,0),  NAVY),
                ("TEXTCOLOR",     (0,0), (-1,0),  WHITE),
                ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,-1), 9),
                ("ALIGN",         (0,0), (-1,0),  "CENTER"),
                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
                ("TEXTCOLOR",     (0,1), (-1,-1), GRAY_DK),
                ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, GRAY_LT]),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("LEFTPADDING",   (0,0), (-1,-1), 8),
                ("RIGHTPADDING",  (0,0), (-1,-1), 8),
                ("GRID",          (0,0), (-1,-1), 0.5, GRAY_BD),
                ("LINEBELOW",     (0,0), (-1,0),  1.5, BLUE),
            ]
            if extra:
                base.extend(extra)
            t.setStyle(TableStyle(base))
            return t

        def section(title):
            return [
                Spacer(1, 4*mm),
                HRFlowable(width="100%", thickness=1.5, color=BLUE, spaceAfter=4),
                Paragraph(title, S_SECTION),
            ]

        story = []

        # Cover
        cover = Table(
            [[Paragraph("AEO AGENT REPORT", ps("ct", fontSize=20, fontName="Helvetica-Bold",
                                                textColor=WHITE, alignment=TA_CENTER))],
             [Paragraph(url, ps("cu", fontSize=11, fontName="Helvetica",
                                textColor=colors.HexColor("#bfdbfe"), alignment=TA_CENTER))],
             [Paragraph(f"Generated: {date}  |  Scrapfly + Z.ai  |  Auto-runs: 1st · 15th · 30th",
                        ps("cd", fontSize=8, fontName="Helvetica",
                           textColor=colors.HexColor("#93c5fd"), alignment=TA_CENTER))]],
            colWidths=[content_w],
        )
        cover.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), NAVY),
            ("TOPPADDING",    (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
            ("LEFTPADDING",   (0,0), (-1,-1), 20),
            ("RIGHTPADDING",  (0,0), (-1,-1), 20),
            ("LINEBELOW",     (0,-1),(-1,-1), 3, BLUE),
        ]))
        story.append(cover)
        story.append(Spacer(1, 6*mm))

        # Score cards
        story.extend(section("AEO VISIBILITY SCORES"))
        score_items = [
            ("Overall AEO",          aeo.get("overall_aeo", 0)),
            ("ChatGPT Visibility",   aeo.get("chatgpt_visibility", 0)),
            ("Gemini Visibility",    aeo.get("gemini_visibility", 0)),
            ("Google AI Overview",   aeo.get("google_ai_overview", 0)),
            ("Content Quality",      aeo.get("content_quality", 0)),
            ("Structured Data",      aeo.get("structured_data", 0)),
            ("Conversational Ready", aeo.get("conversational_readiness", 0)),
            ("Entity Clarity",       aeo.get("entity_clarity", 0)),
        ]
        card_w = content_w / 4 - 2*mm
        gap = 1.5*mm
        rows = [[], []]
        for i, (label, val) in enumerate(score_items):
            card = Table(
                [[Paragraph(label, ps(f"cl{i}", fontSize=8, fontName="Helvetica",
                                      textColor=GRAY_MD, alignment=TA_CENTER))],
                 [Paragraph(str(val), ps(f"cv{i}", fontSize=22, fontName="Helvetica-Bold",
                                         textColor=score_clr(val), alignment=TA_CENTER))],
                 [Paragraph(rating(val), ps(f"cr{i}", fontSize=8, fontName="Helvetica",
                                            textColor=GRAY_MD, alignment=TA_CENTER))]],
                colWidths=[card_w],
            )
            card.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), score_bg(val)),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("BOX",           (0,0), (-1,-1), 1, GRAY_BD),
            ]))
            rows[0 if i < 4 else 1].append(card)

        grid = Table(rows, colWidths=[card_w + gap]*4, rowHeights=[22*mm, 22*mm])
        grid.setStyle(TableStyle([
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",  (0,0), (-1,-1), gap/2),
            ("RIGHTPADDING", (0,0), (-1,-1), gap/2),
            ("TOPPADDING",   (0,0), (-1,-1), gap/2),
            ("BOTTOMPADDING",(0,0), (-1,-1), gap/2),
        ]))
        story.append(grid)

        # Strengths & Weaknesses
        story.extend(section("STRENGTHS & WEAKNESSES"))

        def sw_cell(items, bg, label):
            rows = [[Paragraph(label, ps(f"swh{label}", fontSize=9, fontName="Helvetica-Bold",
                                         textColor=WHITE, alignment=TA_CENTER))]]
            for item in items:
                rows.append([Paragraph(f"{'+ ' if 'STRENGTHS' in label else '- '}{item}",
                                        ps(f"sw{item[:4]}", fontSize=9, fontName="Helvetica",
                                           textColor=GRAY_DK, leading=13))])
            t = Table(rows, colWidths=[content_w/2 - 3*mm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0),  bg),
                ("BACKGROUND", (0,1), (-1,-1), WHITE),
                ("TOPPADDING", (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0),(-1,-1),5),
                ("LEFTPADDING",(0,0), (-1,-1), 8),
                ("RIGHTPADDING",(0,0),(-1,-1), 8),
                ("BOX",        (0,0), (-1,-1), 1, GRAY_BD),
                ("LINEBELOW",  (0,0), (-1,0),  2, bg),
            ]))
            return t

        sw = Table(
            [[sw_cell(aeo.get("strengths", []),  GREEN, "STRENGTHS"),
              sw_cell(aeo.get("weaknesses", []), RED,   "WEAKNESSES")]],
            colWidths=[content_w/2 - 1*mm, content_w/2 - 1*mm],
        )
        sw.setStyle(TableStyle([
            ("LEFTPADDING",  (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
            ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ]))
        story.append(sw)

        # Competitors
        comps = report.get("competitors", [])
        if comps:
            story.extend(section("COMPETITOR INTELLIGENCE"))
            comp_rows = [["Competitor", "Score", "Their Strengths", "Their Weaknesses", "Opportunity"]]
            comp_extras = []
            for i, c in enumerate(comps, 1):
                v = c.get("their_aeo_score", 0)
                comp_rows.append([
                    c.get("competitor_name", "") or c.get("competitor_url", "")[:20],
                    str(v),
                    "\n".join(f"+ {s}" for s in (c.get("strengths", []) or [])[:2]),
                    "\n".join(f"- {w}" for w in (c.get("weaknesses", []) or [])[:2]),
                    c.get("opportunity", "") or "",
                ])
                comp_extras += [
                    ("TEXTCOLOR", (1,i), (1,i), score_clr(v)),
                    ("FONTNAME",  (1,i), (1,i), "Helvetica-Bold"),
                    ("ALIGN",     (1,i), (1,i), "CENTER"),
                ]
            story.append(make_table(comp_rows, [30*mm, 14*mm, 38*mm, 38*mm, 50*mm], comp_extras))

        # Recommendations
        recs = ins.get("recommendations", [])
        if recs:
            story.extend(section("STRATEGIC RECOMMENDATIONS"))
            rec_rows = [["Priority", "Category", "Action", "Impact"]]
            rec_extras = []
            for i, r in enumerate(recs, 1):
                p = r.get("priority", "")
                pcol = RED if p == "HIGH" else (AMBER if p == "MEDIUM" else GREEN)
                pbg = RED_LT if p == "HIGH" else (AMBER_LT if p == "MEDIUM" else GREEN_LT)
                rec_rows.append([p, r.get("category", ""), r.get("action", ""), r.get("impact", "")])
                rec_extras += [
                    ("TEXTCOLOR",  (0,i), (0,i), pcol),
                    ("BACKGROUND", (0,i), (0,i), pbg),
                    ("FONTNAME",   (0,i), (0,i), "Helvetica-Bold"),
                    ("ALIGN",      (0,i), (0,i), "CENTER"),
                ]
            story.append(make_table(rec_rows, [18*mm, 35*mm, 65*mm, 52*mm], rec_extras))

        # FAQs
        faqs = ins.get("trending_faqs", [])
        if faqs:
            story.extend(section("TRENDING FAQs TO ADD TO YOUR SITE"))
            for i, f in enumerate(faqs, 1):
                platform = f.get("ai_platform", "Both")
                plat_col = BLUE if "ChatGPT" in platform else (GREEN if "Gemini" in platform else AMBER)
                q_tbl = Table(
                    [[Paragraph(f"Q{i}: {f.get('question', '')}", S_FAQ_Q),
                      Paragraph(platform, ps(f"fp{i}", fontSize=8, fontName="Helvetica-Bold",
                                             textColor=WHITE, alignment=TA_CENTER))]],
                    colWidths=[content_w - 22*mm, 20*mm],
                )
                q_tbl.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0),(0,0), BLUE_LT),
                    ("BACKGROUND",    (1,0),(1,0), plat_col),
                    ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
                    ("TOPPADDING",    (0,0),(-1,-1), 5),
                    ("BOTTOMPADDING", (0,0),(-1,-1), 5),
                    ("LEFTPADDING",   (0,0),(-1,-1), 8),
                    ("RIGHTPADDING",  (0,0),(-1,-1), 8),
                    ("BOX",           (0,0),(-1,-1), 0.5, GRAY_BD),
                ]))
                story.append(KeepTogether([q_tbl, Paragraph(f.get("answer", ""), S_FAQ_A)]))
                story.append(Spacer(1, 1*mm))

        # 30-Day Plan
        plan = ins.get("next_30_day_plan", [])
        if plan:
            story.extend(section("30-DAY ACTION PLAN"))
            plan_colors = [BLUE, GREEN, AMBER, NAVY]
            for i, item in enumerate(plan):
                parts = item.split(":", 1)
                week = parts[0].strip()
                task = parts[1].strip() if len(parts) > 1 else item
                row = Table(
                    [[Paragraph(week, ps(f"pw{i}", fontSize=9, fontName="Helvetica-Bold",
                                         textColor=WHITE, alignment=TA_CENTER)),
                      Paragraph(task, ps(f"pt{i}", fontSize=10, fontName="Helvetica",
                                         textColor=GRAY_DK, leading=15))]],
                    colWidths=[24*mm, content_w - 26*mm],
                )
                row.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0),(0,0), plan_colors[i % 4]),
                    ("BACKGROUND",    (1,0),(1,0), WHITE if i % 2 == 0 else GRAY_LT),
                    ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
                    ("TOPPADDING",    (0,0),(-1,-1), 8),
                    ("BOTTOMPADDING", (0,0),(-1,-1), 8),
                    ("LEFTPADDING",   (0,0),(-1,-1), 8),
                    ("RIGHTPADDING",  (0,0),(-1,-1), 8),
                    ("BOX",           (0,0),(-1,-1), 0.5, GRAY_BD),
                ]))
                story.append(row)
                story.append(Spacer(1, 1.5*mm))

        # Footer
        story.append(Spacer(1, 8*mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY_BD))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f"AEO Agent  |  {url}  |  {date}  |  Scrapfly + Z.ai",
            S_FOOTER,
        ))

        doc.build(story)
        buf.seek(0)
        safe = url.replace("https://", "").replace("http://", "").replace("/", "_").replace(".", "_")
        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True,
                         download_name=f"AEO_Report_{safe}_{date}.pdf")

    except Exception as exc:
        logger.error("PDF generation failed: %s", exc)
        logger.error(traceback.format_exc())
        from flask import jsonify
        return jsonify({"error": f"PDF error: {exc}"}), 500
